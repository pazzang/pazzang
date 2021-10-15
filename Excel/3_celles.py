from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보 출력
print(ws["A1"].value) # A1 셀의 '값' 출력
print(ws["A10"].value) # 값이 없으면 'None'

# row = 1, 2, 3, ...
# column = A(1), B(2), C(3), ...
print(ws.cell(column=1, row=1).value) # ws["A1"]
print(ws.cell(column=2, row=1).value) # ws["B1"]

c = ws.cell(column=3, row=1, value=10) # ws["C1"].value
print(c.value)

from random import *

# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1, 11): # 10개 row
    for y in range(1, 11): # 10개 column
        ws.cell(column=x, row=y, value=randint(0,100)) # 0~100 사이의 숫자

#34.09

wb.save("sample.xlsx")