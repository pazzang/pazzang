# from openpyxl import Workbook
# wb = Workbook () # 새 워크북 생성
# ws = wb.active # 현재 활성화된 sheet 가져옴
# ws.title = "Test" # sheet의 이름을 변경
# wb.save("sample.xlsx")
# wb.close()


#################################################################################


from openpyxl import Workbook
wb = Workbook ()

ws = wb.create_sheet () # 새로운 Sheet 기본 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번 째 index에 생성
new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근
print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Testtt"
target = wb.copy_worksheet(new_ws)
target.title = "copied Sheet"
target["A2"] = "Copied"

wb.save("sample.xlsx")
wb.close()